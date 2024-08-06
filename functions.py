import re
import pandas as pd
from datetime import datetime
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


def select_files(files_list, date_from, date_to):
    """This function filters the list of filenames according to the start date and the end date of analysis.
    :param files_list: list of files
    :type files_list: list[str]
    :param date_from: start date
    :type date_from: datetime.datetime
    :param date_to: end date
    :type date_to: datetime.datetime
    :return: filtered list of files
    :return type: list[str]
    """
    filtered_files_list = files_list
    temp_date = [datetime.strptime(elem[:-4], '%d.%m.%Y') for elem in filtered_files_list]

    # mark files before analysis start date
    if date_from:
        for i in range(0, len(filtered_files_list)):
            if temp_date[i] < date_from:
                filtered_files_list[i] = ''
    # mark files after analysis end date
    if date_to:
        for i in range(0, len(filtered_files_list)):
            if temp_date[i] > date_to:
                filtered_files_list[i] = ''
    # delete marked items
    filtered_files_list = [elem for elem in filtered_files_list if elem != '']
    return filtered_files_list


def create_dataset(file_paths, files_list):
    """This function read all files and returns a pandas.DataFrame with all data from all selected files.
    :param file_paths: list of paths (PosixPath) to files
    :type file_paths: list[PosixPath]
    :param files_list: list of names of files
    :type files_list: list[str]
    :return: combined data from all datasets
    :return type: pandas.DataFrame
    """
    # sensor_data - pd.DataFrame for all read data
    sensor_data = pd.DataFrame(columns=['name', 'description', 'time', 'alarm', 'date'])

    for i, elem in enumerate(file_paths):
        temp_data = pd.read_csv(elem, names=['name', 'description', 'time', 'alarm'], header=0)
        date = datetime.strptime(files_list[i][:-4], '%d.%m.%Y')
        temp_data['date'] = pd.Series([date] * temp_data.shape[0])
        sensor_data = pd.concat([sensor_data, temp_data])

    return sensor_data


def get_sensor_status(meas_data):
    """This function returns a pandas.DataFrame with sensor response symbol: '-' if measurement time too short (below
half of the median), '+' if proper measurement time and no alarm, '!' if proper measurement time and alarm.
    :param meas_data: DataFrame with combined all measurement results
    :type meas_data: pandas.DataFrame
    :return: modified pandas.DataFrame with columns: 'name', 'date', 'result'
    """
    # ------ data processing ------
    # get the threshold time (half the median of all measurement times)
    time_threshold = 0.5 * np.average(meas_data.loc[:, 'time'])
    # results: '-' if time < time_threshold, '+' if time >= time_threshold, '!' if time >= time_threshold and alarm=T
    result = []
    for i in range(0, meas_data.shape[0]):
        if meas_data.iloc[i, 2] < time_threshold:
            result.append('-')
        elif meas_data.iloc[i, 2] >= time_threshold and meas_data.iloc[i, 3] == 'No':
            result.append('+')
        elif meas_data.iloc[i, 2] >= time_threshold and meas_data.iloc[i, 3] == 'Yes':
            result.append('!')
    meas_data['result'] = result
    meas_data = meas_data.drop(labels=['description', 'time', 'alarm'], axis=1)
    return meas_data.reset_index(drop=True)


def get_result_data(sensor_data):
    """This function process concatenated data from measurement files and transform them into a pandas.DataFrame. The
structure of the output data is as follows: each row corresponds to a unique sensor, each column correspond to a
measurement date, '+' indicates that measurement was correct (not too short and no alarm on), '-' indicates that the
measurement was incorrect (too short), '!'indicates that the measurement was correct (not too short), but the alarm was
on.
    :param sensor_data:
    :type sensor_data: pandas.DataFrame
    :return: processed data
    """
    # creating a new pd.DataFrame for results with '-' values
    rows_names = sensor_data['name'].unique()
    columns_names = sensor_data['date'].unique()
    dummy_data = [['-'] * len(columns_names)] * len(rows_names)
    result_data = pd.DataFrame(data=dummy_data, columns=columns_names, index=rows_names)
    # index for sorting
    result_data['No'] = [int(re.search(r'[0-9]+', elem).group(0)) for elem in rows_names]

    # results reshaping
    for i in range(0, sensor_data.shape[0]):
        result_data.loc[sensor_data.iloc[i, 0], sensor_data.iloc[i, 1]] = sensor_data.iloc[i, 2]

    # sort by sensor index
    result_data = result_data.sort_values(by='No')
    # drop 'No' column
    result_data.drop('No', axis=1,  inplace=True)
    # sort columns - ascending measurement date
    result_data = result_data[result_data.columns.sort_values()]
    # change columns names to format 'dd/mm/yyyy'
    new_col_names = [elem.strftime('%d/%m/%Y') for elem in result_data.columns]
    new_col_dict = dict([(result_data.columns[i], new_col_names[i]) for i in range(0, len(new_col_names))])
    result_data.rename(columns=new_col_dict, inplace=True)
    return result_data


def generate_filename(files_list):
    """This function generated the filename of the  output Excel file in the format: results_[date1]_..._[dateN].xlsx
    :param files_list: a list with names of files
    :type files_list: list[str]
    :return: a string with the file name
    """
    filename = 'results'
    # convert str into datetime
    temp_date = [datetime.strptime(elem[:-4], '%d.%m.%Y') for elem in files_list]
    # sort dates (ascending)
    temp_date.sort()
    # convert datetime into str
    ordered_files_list = [elem.strftime('%d.%m.%Y') for elem in temp_date]
    print(ordered_files_list)
    # create filename
    for name in ordered_files_list:
        filename = filename + '_' + name + '_'
    return filename[:-1] + '.xlsx'


def format_excel_file(filename, r, c):
    """This function opens the Excel file with the results, formats it, and saves changes.
    :param filename: name of the xlsx file
    :type filename: str
    :param r: number of rows in the result table
    :type r: int
    :param c: number of columns in the result table
    :type c: int
    :return: None
    """
    wb = load_workbook(filename=filename)
    my_sheet = wb['Arkusz1']
    my_sheet.cell(2, 2).value = 'name'
    # format table body
    for i in range(3, r + 3):
        for j in range(3, c + 3):
            current_cell = my_sheet.cell(i, j)
            current_cell.alignment = Alignment(horizontal='center')
            current_cell.font = Font(name='Calibri', size=11)
            current_cell.border = Border(left=Side(border_style='medium', color='FF999999'),
                                         right=Side(border_style='medium', color='FF999999'),
                                         top=Side(border_style='medium', color='FF999999'),
                                         bottom=Side(border_style='medium', color='FF999999'))
            if current_cell.value == '+':
                current_cell.font = Font(color='FF006100')
                current_cell.fill = PatternFill("solid", fgColor="FFC6EFCE")
            elif current_cell.value == '-':
                current_cell.font = Font(color='FF9C0006')
                current_cell.fill = PatternFill("solid", fgColor="FFFFC7CE")
            else:
                current_cell.font = Font(color='FF9C6500')
                current_cell.fill = PatternFill("solid", fgColor="FFFFEB9C")

    # format 1st column
    for i in range(2, r + 3):
        current_cell = my_sheet.cell(i, 2)
        current_cell.alignment = Alignment(horizontal='left')
        current_cell.font = Font(name='Liberation Sans', size=10)
        current_cell.border = Border(left=Side(border_style='medium', color='FF999999'),
                                     right=Side(border_style='medium', color='FF999999'),
                                     top=Side(border_style='medium', color='FF999999'),
                                     bottom=Side(border_style='medium', color='FF999999'))
    # format 1st row
    for i in range(2, c + 3):
        current_cell = my_sheet.cell(2, i)
        current_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_cell.font = Font(name='Liberation Sans', size=7)
        current_cell.border = Border(left=Side(border_style='medium', color='FF999999'),
                                     right=Side(border_style='medium', color='FF999999'),
                                     top=Side(border_style='medium', color='FF999999'),
                                     bottom=Side(border_style='double', color='FF999999'))
        current_cell.fill = PatternFill("solid", fgColor="FFF5F5F5")
    wb.save(filename)
